from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook

from fai_template_writer import fill_fai_template, template_row_capacity


@dataclass
class FakeCharacteristic:
    char_number: int
    reference_location: str
    nominal: float
    lsl: float
    usl: float
    type: str
    tooling: str
    comments: str
    actual: str = ""
    metadata: dict = field(default_factory=dict)


def make_r3_like_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FAI FORM"
    ws["A3"] = "FIRST ARTICLE INSPECTION (FAI)"
    ws["A6"] = "Part No:"
    ws["J6"] = "Part Name:"
    ws["A8"] = "Date:"
    ws["J8"] = "Drawing No:"
    ws["A10"] = "Insp. by:"
    ws["J10"] = "Revision:"
    ws["A12"] = "Item No:"
    ws["J12"] = "PO No:"
    ws["A14"] = "Order No:"
    ws["J14"] = "Reason for FAI:"
    ws["A22"] = "CHAR"
    ws["C22"] = "REQUIREMENT"
    ws["G22"] = "SUPPLIER INSPECTION RESULT"
    ws["J22"] = "EZ FABRICATING INSPECTION RESULT"
    for merge_ref in ["B6:E6", "K6:N6", "B8:E8", "K8:N8", "B10:E10", "K10:N10", "B12:E12", "K12:N12", "B14:E14", "K14:N14"]:
        ws.merge_cells(merge_ref)
    for row in range(24, 49):
        ws.cell(row=row, column=1).value = row - 23
        ws.cell(row=row, column=11).value = f'=IF(J{row}="","",IF(AND(J{row}>=C{row},J{row}<=E{row}),"X",""))'
    wb.create_sheet("CHARACTERISTICS")
    wb.create_sheet("TOOLING")
    wb.create_sheet("ATTRIBUTE")
    wb.save(path)


def test_r3_template_writer_preserves_form_and_inclusive_formula(tmp_path):
    template = tmp_path / "EZ_FAB_1st_Article_Form_R3.xlsx"
    output = tmp_path / "filled.xlsx"
    make_r3_like_template(template)
    chars = [
        FakeCharacteristic(1, "P1-R1C1", 16.0, 15.98, 16.02, "LINEAR", "CALIPER", "AFTER GALVANIZE"),
        FakeCharacteristic(2, "P1-R1C2", 0.81, 0.805, 0.815, "Ø", "CALIPER", ""),
    ]

    assert template_row_capacity(template) == 25
    fill_fai_template(template, chars, output)

    wb = load_workbook(output, data_only=False)
    ws = wb["FAI FORM"]
    assert ws["A3"].value == "FIRST ARTICLE INSPECTION (FAI)"
    assert ws["B24"].value == "P1-R1C1"
    assert ws["C24"].value == 15.98
    assert ws["D24"].value == 16.0
    assert ws["E24"].value == 16.02
    assert ws["F24"].value == "LINEAR"
    assert ws["J24"].value is None
    assert ">=C24" in ws["K24"].value
    assert "<=E24" in ws["K24"].value
    assert ws["M24"].value == "CALIPER"
    assert ws["N24"].value == "AFTER GALVANIZE"
    assert ws.row_dimensions[24].height >= 14
    assert ws.print_area == "'FAI FORM'!$A$1:$N$48"
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 1
    assert len(ws.data_validations.dataValidation) >= 3


def test_r3_template_writer_fills_only_non_admin_header_metadata(tmp_path):
    template = tmp_path / "EZ_FAB_1st_Article_Form_R3.xlsx"
    output = tmp_path / "filled.xlsx"
    make_r3_like_template(template)
    chars = [
        FakeCharacteristic(
            1,
            "P1-R1C1",
            16.0,
            15.98,
            16.02,
            "LINEAR",
            "CALIPER",
            "",
            metadata={
                "part_no": "V9050SP-104B-01",
                "part_name": "LOUVER INTAKE REAR SIDE CAP",
                "drawing_no": "V9050SP-104B-01",
                "revision": "A",
                "date": "SHOULD NOT WRITE",
                "inspector": "SHOULD NOT WRITE",
                "item_no": "SHOULD NOT WRITE",
                "order_no": "SHOULD NOT WRITE",
                "po_no": "SHOULD NOT WRITE",
                "reason_for_fai": "SHOULD NOT WRITE",
            },
        )
    ]

    fill_fai_template(template, chars, output)
    wb = load_workbook(output, data_only=False)
    ws = wb["FAI FORM"]
    assert ws["B6"].value == "V9050SP-104B-01"
    assert ws["K6"].value == "LOUVER INTAKE REAR SIDE CAP"
    assert ws["K8"].value == "V9050SP-104B-01"
    assert ws["K10"].value == "A"
    assert ws["B8"].value is None
    assert ws["B10"].value is None
    assert ws["B12"].value is None
    assert ws["B14"].value is None
    assert ws["K12"].value is None
    assert ws["K14"].value is None


def test_generic_template_still_works(tmp_path):
    template = tmp_path / "generic.xlsx"
    output = tmp_path / "generic_filled.xlsx"
    wb = Workbook()
    ws = wb.active
    headers = ["Char Number", "Reference Location", "Requirement LSL", "Requirement Nominal", "Requirement USL", "Type", "EZ Fabricating Actual", "In Spec", "Tooling Used", "Comments"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    wb.save(template)

    fill_fai_template(template, [FakeCharacteristic(1, "P1", 4.69, 4.66, 4.82, "LINEAR", "CALIPER", "")], output)
    wb = load_workbook(output, data_only=False)
    ws = wb.active
    assert ws["B2"].value == "P1"
    assert ">=C2" in ws["H2"].value
    assert "<=E2" in ws["H2"].value
