from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from built_in_form_writer import write_inspection_workbook
from form_profiles import build_default_configuration
from project_store import ProjectMetadata, ProjectRecord, ProjectStore


def test_project_store_round_trip(tmp_path: Path) -> None:
    store = ProjectStore(tmp_path / "projects.db")
    project = ProjectRecord(
        name="A100 Rev B",
        source_pdf="C:/drawings/A100.pdf",
        metadata=ProjectMetadata(part_no="A100", drawing_no="A100", revision="B"),
        characteristics=[{"char_number": 1, "nominal": 1.0}],
    )
    store.save(project)
    loaded = store.load(project.id)
    assert loaded.name == "A100 Rev B"
    assert loaded.metadata.part_no == "A100"
    assert loaded.characteristics[0]["char_number"] == 1
    assert store.recent(1)[0]["id"] == project.id


def test_builtin_workbook_contains_required_fields_and_editable_tooling(tmp_path: Path) -> None:
    output = tmp_path / "inspection.xlsx"
    rows = [{
        "char_number": 1,
        "reference_location": "P1-R2C3",
        "lsl": 0.995,
        "nominal": 1.0,
        "usl": 1.005,
        "type": "DIAMETER",
        "actual": 1.001,
        "tooling": "MICROMETER [M-17]",
        "comments": "Verified",
    }]
    write_inspection_workbook(
        output,
        ProjectMetadata(part_no="A100", drawing_no="A100", revision="B"),
        rows,
        build_default_configuration(),
    )
    workbook = load_workbook(output)
    sheet = workbook["Inspection Report"]
    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert "EZ FAIR INSPECTION REPORT" in values
    assert "Part No." in values
    assert "Qualified Tooling" in values
    assert "MICROMETER [M-17]" in values
    assert sheet.data_validations.count >= 1
