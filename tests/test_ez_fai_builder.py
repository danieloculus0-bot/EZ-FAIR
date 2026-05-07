from pathlib import Path

import fitz
from openpyxl import Workbook, load_workbook

from ez_fai_builder import (
    Characteristic,
    add_pdf_balloons,
    calculate_tolerance_limits,
    classify_dimension,
    extract_pdf_dimensions,
    fill_fai_template,
    get_last_skipped_candidates,
    write_debug_report,
)


def test_classify_dimension():
    assert classify_dimension("Ø .97") == "Ø"
    assert classify_dimension("76.00°") == "°"
    assert classify_dimension("FILLET WELD .25") == "WELD"
    assert classify_dimension("16.00") == "LINEAR"


def test_tolerance_logic_defaults_and_explicit():
    assert calculate_tolerance_limits(16.00, "16.00", "LINEAR", "") == (15.98, 16.02)
    assert calculate_tolerance_limits(0.970, ".970", "Ø", "") == (0.965, 0.975)
    assert calculate_tolerance_limits(76.00, "76.00", "°", "") == (74.0, 78.0)
    assert calculate_tolerance_limits(5.52, "5.52", "LINEAR", "5.52 +.13 / -.03") == (5.49, 5.65)


def _make_template(path: Path, headers: list[str] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    for idx, header in enumerate(headers or [
        "Char Number",
        "Reference Location",
        "Requirement LSL",
        "Requirement Nominal",
        "Requirement USL",
        "Type",
        "EZ Fabricating Actual",
        "In Spec",
        "Tooling Used",
        "Comments",
    ], start=1):
        ws.cell(row=1, column=idx).value = header
    wb.save(path)


def test_fill_fai_template_writes_inclusive_formula(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _make_template(template)

    characteristic = Characteristic(
        char_number=1,
        reference_location="P1-R1C1",
        nominal=16.0,
        lsl=15.98,
        usl=16.02,
        type="LINEAR",
        page_index=0,
        rect=(0, 0, 10, 10),
        tooling="CALIPER",
    )
    output = fill_fai_template(template, [characteristic], tmp_path / "drawing_FAI.xlsx")

    result = load_workbook(output, data_only=False)
    ws = result.active
    assert ws["A2"].value == 1
    assert ws["H2"].value == '=IF(G2="","",IF(AND(G2>=C2,G2<=E2),"X",""))'


def test_fill_fai_template_uses_alias_headers_and_leaves_admin_blank(tmp_path: Path):
    template = tmp_path / "corrected_template.xlsx"
    _make_template(
        template,
        [
            "PO",
            "Order Number",
            "Characteristic No.",
            "Location",
            "Lower Spec Limit",
            "Nominal",
            "Upper Spec Limit",
            "Actual Result",
            "Pass/Fail",
            "Inspection Tool",
            "Remarks",
        ],
    )
    characteristic = Characteristic(1, "P1-R1C1", 5.52, 5.49, 5.65, "LINEAR", 0, (0, 0, 10, 10), tooling="CALIPER")

    output = fill_fai_template(template, [characteristic], tmp_path / "drawing_FAI.xlsx")

    ws = load_workbook(output, data_only=False).active
    assert ws["A2"].value is None
    assert ws["B2"].value is None
    assert ws["C2"].value == 1
    assert ws["I2"].value == '=IF(H2="","",IF(AND(H2>=E2,H2<=G2),"X",""))'


def test_extract_balloon_and_debug_report_end_to_end(tmp_path: Path):
    pdf = tmp_path / "DVM-AE.pdf"
    doc = fitz.open()
    page = doc.new_page(width=400, height=300)
    page.insert_text((60, 60), "16.00  76.00°  Ø .970")
    page.insert_text((60, 90), "5.52 +.13 / -.03")
    page.insert_text((260, 250), "DRAWING NO 1234.56 REV A")
    page.insert_text((60, 120), "NOTE: GALVANIZE AFTER FAB")
    doc.save(pdf)
    doc.close()

    chars = extract_pdf_dimensions(pdf)
    assert len(chars) >= 4
    linear_16 = next(c for c in chars if c.raw_text.startswith("16.00"))
    assert linear_16.tooling == "CALIPER"
    assert (linear_16.lsl, linear_16.usl) == (15.98, 16.02)
    assert any(c.type == "°" and c.tooling == "ANGLE GAGE" and (c.lsl, c.usl) == (74.0, 78.0) for c in chars)
    assert any(c.type == "Ø" for c in chars)
    assert all(c.comments == "AFTER GALVANIZE" for c in chars)
    assert any(s.reason for s in get_last_skipped_candidates())

    ballooned = add_pdf_balloons(pdf, chars, tmp_path / "DVM-AE_BALLOONED.pdf")
    assert ballooned.exists()

    template = tmp_path / "template.xlsx"
    _make_template(template)
    fai = fill_fai_template(template, chars, tmp_path / "DVM-AE_FAI.xlsx")
    assert fai.exists()

    report = write_debug_report(pdf, template, chars, tmp_path / "EZ_FAI_DEBUG_REPORT.txt")
    text = report.read_text(encoding="utf-8")
    assert "PDF file used" in text
    assert "Number of extracted characteristics" in text
    assert "Skipped dimension candidates" in text


def test_local_paths_and_customer_files_are_ignored():
    gitignore = Path(".gitignore").read_text(encoding="utf-8").splitlines()
    assert "local_inputs/" in gitignore
    assert "local_outputs/" in gitignore
    assert "*.pdf" in gitignore
    assert "*.xlsx" in gitignore
    assert "*.xlsm" in gitignore
    assert "*.py" not in gitignore


def test_local_test_runner_creates_expected_outputs_and_summary(tmp_path: Path):
    from local_test_runner import run_local_test

    input_dir = tmp_path / "local_inputs"
    output_dir = tmp_path / "local_outputs"
    input_dir.mkdir()

    pdf = input_dir / "DVM-AE.pdf"
    doc = fitz.open()
    page = doc.new_page(width=400, height=300)
    page.insert_text((60, 60), "16.00  76.00°  Ø .970")
    page.insert_text((60, 90), "5.52 +.13 / -.03")
    page.insert_text((260, 250), "DRAWING NO 1234.56 REV A")
    doc.save(pdf)
    doc.close()

    template = input_dir / "corrected_template.xlsx"
    _make_template(
        template,
        [
            "PO",
            "Order Number",
            "Characteristic No.",
            "Location",
            "Lower Spec Limit",
            "Nominal",
            "Upper Spec Limit",
            "Actual Result",
            "Pass/Fail",
            "Inspection Tool",
            "Remarks",
        ],
    )

    outputs = run_local_test(pdf, template, output_dir)

    assert outputs["ballooned_pdf"] == output_dir / "DVM-AE_BALLOONED.pdf"
    assert outputs["fai_excel"] == output_dir / "DVM-AE_FAI.xlsx"
    assert outputs["debug_report"] == output_dir / "EZ_FAI_DEBUG_REPORT.txt"
    assert outputs["extraction_summary"] == output_dir / "EXTRACTION_SUMMARY.txt"
    assert outputs["ballooned_pdf"].exists()
    assert outputs["fai_excel"].exists()
    assert outputs["debug_report"].exists()
    assert outputs["extraction_summary"].exists()

    summary = outputs["extraction_summary"].read_text(encoding="utf-8")
    assert "Total extracted characteristics" in summary
    assert "Count by type" in summary
    assert "Count of skipped candidates" in summary
    assert "Likely false positives" in summary
    assert "Likely duplicate dimensions" in summary
    assert "Dimensions with missing or guessed tolerance" in summary
    assert "Dimensions with explicit tolerance" in summary
    assert "Title block tolerance defaults used" in summary
