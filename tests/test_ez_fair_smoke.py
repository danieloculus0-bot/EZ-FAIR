from pathlib import Path

import fitz
from openpyxl import Workbook, load_workbook

from ez_fai_builder import calculate_tolerance_limits, extract_pdf_dimensions
from local_test_runner import run_local_test


def make_demo_pdf(path: Path) -> None:
    doc = fitz.open()
    page = doc.new_page(width=792, height=612)
    page.insert_text((72, 72), "DVM-AE DEMO DRAWING", fontsize=10)
    page.insert_text((72, 120), "16.00", fontsize=12)
    page.insert_text((140, 120), "76.00°", fontsize=12)
    page.insert_text((220, 120), "Ø .810", fontsize=12)
    page.insert_text((72, 170), "4.69 +.13 -.03", fontsize=12)
    page.insert_text((72, 220), "NO WELD AT SLOT LOCATIONS .25", fontsize=12)
    page.insert_text((72, 420), "DIMENSIONS ARE IN INCHES", fontsize=9)
    page.insert_text((72, 440), "TWO PLACE DECIMAL 0.02", fontsize=9)
    page.insert_text((72, 460), "THREE PLACE DECIMAL 0.005", fontsize=9)
    page.insert_text((72, 480), "ANGULAR 2", fontsize=9)
    page.insert_text((72, 500), "ALL METAL PARTS TO BE GALVANIZED", fontsize=9)
    doc.save(path)
    doc.close()


def make_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    headers = [
        "Char Number",
        "Reference Location",
        "Requirement LSL",
        "Requirement Nominal",
        "Requirement USL",
        "Type",
        "EZ Fabricating Actual",
        "In Spec",
        "Tooling Used",
        "Comments",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    wb.save(path)


def test_tolerance_limits_use_expected_bounds():
    assert calculate_tolerance_limits(16.00, "16.00", "LINEAR") == (15.98, 16.02)
    assert calculate_tolerance_limits(0.810, ".810", "Ø") == (0.805, 0.815)
    assert calculate_tolerance_limits(76.00, "76.00", "°") == (74.0, 78.0)
    assert calculate_tolerance_limits(4.69, "4.69", "LINEAR", "4.69 +.13 -.03") == (4.66, 4.82)


def test_demo_pdf_extracts_expected_characteristics(tmp_path):
    pdf = tmp_path / "DVM-AE.pdf"
    make_demo_pdf(pdf)

    chars = extract_pdf_dimensions(pdf)
    assert len(chars) >= 4
    assert any(c.type == "°" for c in chars)
    assert any(c.type == "Ø" for c in chars)
    assert any(c.type == "WELD" for c in chars)
    assert any(c.comments == "AFTER GALVANIZE" for c in chars)


def test_local_runner_creates_artifacts_and_inclusive_formula(tmp_path):
    pdf = tmp_path / "DVM-AE.pdf"
    template = tmp_path / "EZ_FAB_FAI_TEMPLATE.xlsx"
    out = tmp_path / "out"
    make_demo_pdf(pdf)
    make_template(template)

    outputs = run_local_test(pdf, template, out)

    assert outputs["characteristic_count"] >= 4
    assert (out / "DVM-AE_BALLOONED.pdf").exists()
    assert (out / "DVM-AE_FAI.xlsx").exists()
    assert (out / "EZ_FAI_DEBUG_REPORT.txt").exists()
    assert (out / "EXTRACTION_SUMMARY.txt").exists()

    wb = load_workbook(out / "DVM-AE_FAI.xlsx", data_only=False)
    ws = wb.active
    formula = ws["H2"].value
    assert ">=C2" in formula
    assert "<=E2" in formula
