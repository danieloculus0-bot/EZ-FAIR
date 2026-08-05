from __future__ import annotations

import copy
import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

EXACT_R3_SHEET = "FAI FORM"
R3_START_ROW = 24
R3_TEMPLATE_END_ROW = 48
R3_TEMPLATE_CAPACITY = 25
R3_COLUMNS = {
    "Char Number": 1, "Reference Location": 2, "Requirement LSL": 3,
    "Requirement Nominal": 4, "Requirement USL": 5, "Type": 6,
    "Supplier Actual": 7, "Supplier Yes": 8, "Supplier No": 9,
    "EZ Fabricating Actual": 10, "In Spec": 11, "Tooling Used": 13, "Comments": 14,
}
GENERIC_HEADERS = [
    "Char Number", "Reference Location", "Requirement LSL", "Requirement Nominal",
    "Requirement USL", "Type", "EZ Fabricating Actual", "In Spec", "Tooling Used", "Comments",
]
HEADER_ALIASES = {
    "Char Number": ["char", "char number", "char no", "characteristic", "balloon"],
    "Reference Location": ["reference location", "ref location", "location", "zone"],
    "Requirement LSL": ["requirement lsl", "lsl", "lower limit", "minimum", "min"],
    "Requirement Nominal": ["requirement nominal", "nominal", "requirement", "dimension"],
    "Requirement USL": ["requirement usl", "usl", "upper limit", "maximum", "max"],
    "Type": ["type", "characteristic type"],
    "EZ Fabricating Actual": ["ez fabricating actual", "actual", "measurement"],
    "In Spec": ["in spec", "pass fail", "accept", "result"],
    "Tooling Used": ["tooling used", "tooling", "gage", "inspection tool"],
    "Comments": ["comments", "notes", "remarks"],
}
LIST_SHEET_SPECS = {
    "CHARACTERISTICS": ["NOTE", "TOLERANCE", "FINISH", "WELD", "MATERIAL", "RADIUS", "LINEAR", "DIAMETER", "Ø", "ANGLE", "THREAD"],
    "TOOLING": ["VISUAL", "CALIPER", "CERTIFICATION", "TAPE", "PROTRACTOR", "ANGLE GAGE", "THREAD GAGE", "HARDWARE", "FITMENT/NHA", "MICROMETER"],
    "ATTRIBUTE": ["PASS", "FAIL"],
}


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _safe(characteristic: Any, attr: str, default: Any = "") -> Any:
    return getattr(characteristic, attr, default)


def _metadata(characteristic: Any) -> dict[str, Any]:
    value = getattr(characteristic, "metadata", {}) or {}
    return value if isinstance(value, dict) else {}


def _row_values(characteristic: Any) -> dict[str, Any]:
    return {
        "Char Number": _safe(characteristic, "char_number"),
        "Reference Location": _safe(characteristic, "reference_location"),
        "Requirement LSL": _safe(characteristic, "lsl"),
        "Requirement Nominal": _safe(characteristic, "nominal"),
        "Requirement USL": _safe(characteristic, "usl"),
        "Type": _safe(characteristic, "type"),
        "EZ Fabricating Actual": _safe(characteristic, "actual", ""),
        "Tooling Used": _safe(characteristic, "tooling", ""),
        "Comments": _safe(characteristic, "comments", ""),
    }


def _pick_sheet(workbook):
    return workbook[EXACT_R3_SHEET] if EXACT_R3_SHEET in workbook.sheetnames else workbook.active


def _is_r3_form(sheet) -> bool:
    markers = " ".join(str(sheet[cell].value or "").upper() for cell in ("A3", "A22", "C22", "G22", "J22"))
    return "FIRST ARTICLE" in markers and "CHAR" in markers and "REQUIREMENT" in markers and "INSPECTION RESULT" in markers


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy.copy(source.alignment)
        target.border = copy.copy(source.border)
        target.fill = copy.copy(source.fill)
        target.font = copy.copy(source.font)
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height or 17


def _ensure_lists(workbook) -> None:
    for name, values in LIST_SHEET_SPECS.items():
        sheet = workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)
        for row, value in enumerate(values, 1):
            sheet.cell(row, 1, value)


def _add_validation(sheet, range_ref: str, formula: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(range_ref)


def _inclusive_formula(row: int) -> str:
    return f'=IF(J{row}="","",IF(AND(J{row}>=C{row},J{row}<=E{row}),"X",""))'


def _fill_r3_metadata(sheet, characteristics: list[Any]) -> None:
    if not characteristics:
        return
    metadata = _metadata(characteristics[0])
    approved = {
        "part_no": "B6",
        "part_name": "K6",
        "drawing_no": "K8",
        "revision": "K10",
    }
    for key, cell in approved.items():
        if metadata.get(key) not in (None, ""):
            sheet[cell] = metadata[key]


def _ensure_r3_capacity(sheet, count: int) -> int:
    end_row = max(R3_TEMPLATE_END_ROW, R3_START_ROW + max(count, 1) - 1)
    if end_row > R3_TEMPLATE_END_ROW:
        sheet.insert_rows(R3_TEMPLATE_END_ROW + 1, end_row - R3_TEMPLATE_END_ROW)
        for row in range(R3_TEMPLATE_END_ROW + 1, end_row + 1):
            _copy_row_style(sheet, R3_TEMPLATE_END_ROW, row)
    return end_row


def _fill_r3(sheet, characteristics: list[Any]) -> None:
    end_row = _ensure_r3_capacity(sheet, len(characteristics))
    _fill_r3_metadata(sheet, characteristics)
    for row in range(R3_START_ROW, end_row + 1):
        sheet.cell(row, 1, row - R3_START_ROW + 1)
        for key in ["Reference Location", "Requirement LSL", "Requirement Nominal", "Requirement USL", "Type", "Supplier Actual", "Supplier Yes", "Supplier No", "EZ Fabricating Actual", "Tooling Used", "Comments"]:
            sheet.cell(row, R3_COLUMNS[key], None)
        sheet.cell(row, R3_COLUMNS["In Spec"], _inclusive_formula(row))
        sheet.row_dimensions[row].height = 17
    for offset, characteristic in enumerate(characteristics):
        row = R3_START_ROW + offset
        values = _row_values(characteristic)
        for key in ["Char Number", "Reference Location", "Requirement LSL", "Requirement Nominal", "Requirement USL", "Type", "Tooling Used", "Comments"]:
            sheet.cell(row, R3_COLUMNS[key], values[key])
        sheet.cell(row, R3_COLUMNS["EZ Fabricating Actual"], None)
        sheet.cell(row, R3_COLUMNS["In Spec"], _inclusive_formula(row))
    _add_validation(sheet, f"F{R3_START_ROW}:F{end_row}", "'CHARACTERISTICS'!$A$1:$A$11")
    _add_validation(sheet, f"M{R3_START_ROW}:M{end_row}", "'TOOLING'!$A$1:$A$10")
    _add_validation(sheet, f"H{R3_START_ROW}:I{end_row}", "'ATTRIBUTE'!$A$1:$A$2")
    if sheet.sheet_properties.pageSetUpPr is None:
        sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_area = f"A1:N{end_row}"
    sheet.print_title_rows = "1:23"
    sheet.freeze_panes = "A24"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_view.showGridLines = False
    for row in range(R3_START_ROW, end_row + 1):
        for column in range(1, 15):
            cell = sheet.cell(row, column)
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "center", vertical="center", wrap_text=True)


def _header_targets() -> dict[str, str]:
    targets: dict[str, str] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        targets[_norm(canonical)] = canonical
        for alias in aliases:
            targets[_norm(alias)] = canonical
    return targets


def _find_generic_header(sheet) -> tuple[int, dict[str, int]]:
    targets = _header_targets()
    best_row, best = 1, {}
    for row in range(1, min(sheet.max_row, 80) + 1):
        mapping = {}
        for column in range(1, min(sheet.max_column, 80) + 1):
            normalized = _norm(sheet.cell(row, column).value)
            if normalized in targets:
                mapping.setdefault(targets[normalized], column)
        if len(mapping) > len(best):
            best_row, best = row, mapping
    if len(best) < 4:
        best_row = 1
        best = {header: index for index, header in enumerate(GENERIC_HEADERS, 1)}
        for header, column in best.items():
            sheet.cell(best_row, column, header)
    else:
        for header in GENERIC_HEADERS:
            if header not in best:
                column = sheet.max_column + 1
                sheet.cell(best_row, column, header)
                best[header] = column
    return best_row, best


def _fill_generic(sheet, characteristics: list[Any]) -> None:
    header_row, columns = _find_generic_header(sheet)
    for offset, characteristic in enumerate(characteristics):
        row = header_row + 1 + offset
        if offset:
            _copy_row_style(sheet, header_row + 1, row)
        values = _row_values(characteristic)
        for header, value in values.items():
            sheet.cell(row, columns[header], None if header == "EZ Fabricating Actual" else value)
        actual = get_column_letter(columns["EZ Fabricating Actual"])
        lsl = get_column_letter(columns["Requirement LSL"])
        usl = get_column_letter(columns["Requirement USL"])
        sheet.cell(row, columns["In Spec"], f'=IF({actual}{row}="","",IF(AND({actual}{row}>={lsl}{row},{actual}{row}<={usl}{row}),"X",""))')


def template_row_capacity(template_path: str | Path) -> int | None:
    workbook = load_workbook(template_path, read_only=False, data_only=False)
    try:
        sheet = _pick_sheet(workbook)
        return R3_TEMPLATE_CAPACITY if _is_r3_form(sheet) else None
    finally:
        workbook.close()


def fill_fai_template(template_path: str | Path, characteristics: Iterable[Any], output_path: str | Path | None = None) -> Path:
    template = Path(template_path)
    rows = list(characteristics)
    if output_path is None:
        drawing = _metadata(rows[0]).get("drawing_name", "FAI") if rows else "FAI"
        output_path = template.with_name(f"{drawing}_FAI{template.suffix}")
    output = Path(output_path)
    if output.resolve() == template.resolve():
        raise ValueError("FAI Excel output cannot overwrite the original template.")
    workbook = load_workbook(template, keep_vba=template.suffix.lower() == ".xlsm")
    _ensure_lists(workbook)
    sheet = _pick_sheet(workbook)
    if _is_r3_form(sheet):
        _fill_r3(sheet, rows)
    else:
        _fill_generic(sheet, rows)
    workbook.save(output)
    workbook.close()
    return output
