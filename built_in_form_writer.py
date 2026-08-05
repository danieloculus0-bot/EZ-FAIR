"""Generate the approved built-in EZ FAIR inspection workbook."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from form_profiles import FormConfiguration, build_default_configuration, resolve_columns
from project_store import ProjectMetadata

DARK = "15191F"
PANEL = "20262E"
HEADER = "2B333D"
TEXT = "F1F5F9"
MUTED = "AAB4C0"
PASS = "1E6B3A"
FAIL = "8E2D2D"
BORDER = Side(style="thin", color="4B5663")


def _value(item: Any, *names: str, default: Any = "") -> Any:
    if isinstance(item, dict):
        for name in names:
            if name in item:
                return item[name]
        return default
    for name in names:
        if hasattr(item, name):
            return getattr(item, name)
    return default


def _row_value(item: Any, key: str) -> Any:
    mapping = {
        "char_number": ("char_number", "Char Number"),
        "reference_location": ("reference_location", "Reference Location"),
        "lsl": ("lsl", "Requirement LSL"),
        "nominal": ("nominal", "Requirement Nominal"),
        "usl": ("usl", "Requirement USL"),
        "feature_type": ("type", "feature_type", "Type"),
        "supplier_actual": ("supplier_actual",),
        "supplier_result": ("supplier_result",),
        "ez_actual": ("actual", "ez_actual", "EZ Fabricating Actual"),
        "qualified_tooling": ("tooling", "qualified_tooling", "Tooling Used"),
        "comments": ("comments", "Comments"),
    }
    if key == "in_spec":
        return ""
    return _value(item, *mapping.get(key, (key,)))


def write_inspection_workbook(
    output_path: str | Path,
    metadata: ProjectMetadata | dict[str, Any],
    characteristics: Iterable[Any],
    configuration: FormConfiguration | None = None,
) -> Path:
    """Create a stable workbook from approved built-in form components."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    config = configuration or build_default_configuration()
    columns = resolve_columns(config)
    meta = metadata if isinstance(metadata, dict) else metadata.__dict__

    wb = Workbook()
    ws = wb.active
    ws.title = "Inspection Report"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    max_col = max(1, len(columns))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title = ws.cell(1, 1, "EZ FAIR INSPECTION REPORT")
    title.font = Font(size=18, bold=True, color=TEXT)
    title.fill = PatternFill("solid", fgColor=DARK)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    header_fields = [
        ("Part No.", meta.get("part_no", "")),
        ("Part Name", meta.get("part_name", "")),
        ("Drawing No.", meta.get("drawing_no", "")),
        ("Revision", meta.get("revision", "")),
        ("Customer", meta.get("customer", "")),
        ("Material", meta.get("material", "")),
        ("PO No.", meta.get("po_no", "")),
        ("Order No.", meta.get("order_no", "")),
        ("Inspector", meta.get("inspector", "")),
        ("Reason for FAI", meta.get("reason_for_fai", "")),
    ]
    pairs_per_row = max(1, max_col // 4)
    row = 2
    col = 1
    for label, value in header_fields:
        if col + 1 > max_col:
            row += 1
            col = 1
        label_cell = ws.cell(row, col, label)
        value_cell = ws.cell(row, min(col + 1, max_col), value)
        label_cell.font = Font(bold=True, color=TEXT)
        value_cell.font = Font(color=TEXT)
        label_cell.fill = PatternFill("solid", fgColor=HEADER)
        value_cell.fill = PatternFill("solid", fgColor=PANEL)
        label_cell.border = value_cell.border = Border(bottom=BORDER)
        col += 2
    table_header_row = max(7, row + 2)

    for index, column in enumerate(columns, start=1):
        cell = ws.cell(table_header_row, index, column.label)
        cell.font = Font(bold=True, color=TEXT)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
        ws.column_dimensions[get_column_letter(index)].width = column.width

    tooling_values: set[str] = set()
    rows = list(characteristics)
    for offset, item in enumerate(rows, start=1):
        excel_row = table_header_row + offset
        for index, column in enumerate(columns, start=1):
            value = _row_value(item, column.key)
            cell = ws.cell(excel_row, index, value)
            cell.fill = PatternFill("solid", fgColor=PANEL)
            cell.font = Font(color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
            if column.key == "in_spec":
                lookup = {item.key: idx + 1 for idx, item in enumerate(columns)}
                actual_col = lookup.get("ez_actual")
                lsl_col = lookup.get("lsl")
                usl_col = lookup.get("usl")
                if actual_col and lsl_col and usl_col:
                    a = get_column_letter(actual_col)
                    l = get_column_letter(lsl_col)
                    u = get_column_letter(usl_col)
                    cell.value = f'=IF({a}{excel_row}="","",IF(AND({a}{excel_row}>={l}{excel_row},{a}{excel_row}<={u}{excel_row}),"PASS","FAIL"))'
            if column.key == "qualified_tooling" and value:
                tooling_values.add(str(value))

    lists = wb.create_sheet("Approved Lists")
    lists.sheet_state = "hidden"
    default_tools = [
        "VISUAL", "CALIPER", "MICROMETER", "HEIGHT GAGE", "PIN GAGE",
        "THREAD GAGE", "CMM", "SURFACE PLATE", "PROTRACTOR", "ANGLE GAGE",
        "TAPE", "CERTIFICATION", "FITMENT/NHA", "HARDWARE",
    ]
    all_tools = sorted(set(default_tools) | tooling_values)
    for idx, tool in enumerate(all_tools, start=1):
        lists.cell(idx, 1, tool)

    column_lookup = {column.key: index + 1 for index, column in enumerate(columns)}
    tooling_col = column_lookup.get("qualified_tooling")
    if tooling_col:
        dv = DataValidation(type="list", formula1=f"'Approved Lists'!$A$1:$A${max(1, len(all_tools))}", allow_blank=True)
        dv.error = "Choose an approved tool or type a project-specific tool directly."
        dv.errorTitle = "Qualified Tooling"
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(tooling_col)}{table_header_row + 1}:{get_column_letter(tooling_col)}{table_header_row + max(200, len(rows) + 20)}")

    in_spec_col = column_lookup.get("in_spec")
    if in_spec_col:
        letter = get_column_letter(in_spec_col)
        end_row = table_header_row + max(200, len(rows) + 20)
        ws.conditional_formatting.add(
            f"{letter}{table_header_row + 1}:{letter}{end_row}",
            FormulaRule(formula=[f'{letter}{table_header_row + 1}="PASS"'], fill=PatternFill("solid", fgColor=PASS)),
        )
        ws.conditional_formatting.add(
            f"{letter}{table_header_row + 1}:{letter}{end_row}",
            FormulaRule(formula=[f'{letter}{table_header_row + 1}="FAIL"'], fill=PatternFill("solid", fgColor=FAIL)),
        )

    ws.auto_filter.ref = f"A{table_header_row}:{get_column_letter(max_col)}{table_header_row + max(1, len(rows))}"
    ws.print_title_rows = f"1:{table_header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "EZ FAIR | Controlled inspection output"
    wb.save(output)
    return output
