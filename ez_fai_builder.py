"""EZ FAI Builder - local desktop MVP.

Drag/select a PDF blueprint and Excel FAI template, extract likely dimensional
characteristics, review/edit them, then generate a ballooned PDF and filled FAI
workbook without using cloud services, pandas, or a database.
"""
from __future__ import annotations

import argparse
import copy
import re
import tkinter as tk
from dataclasses import dataclass, field
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any, Iterable

import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TITLE_BLOCK_DEFAULTS = {
    "two_place": 0.02,
    "three_place": 0.005,
    "angular": 2.0,
}

DEFAULT_TOOLING = {
    "LINEAR": "CALIPER",
    "Ø": "CALIPER",
    "°": "ANGLE GAGE",
    "WELD": "VISUAL",
}

FAI_HEADERS = [
    "Char Number",
    "Reference Location",
    "Requirement LSL",
    "Requirement Nominal",
    "Requirement USL",
    "Type",
    "EZ Fabricating Actual",
    "In Spec",
    "Tooling Used",
    "Comments",
]


HEADER_ALIASES = {
    "Char Number": ["char", "char no", "char number", "characteristic", "characteristic no", "characteristic number", "balloon", "balloon no"],
    "Reference Location": ["reference location", "ref location", "location", "zone", "drawing zone"],
    "Requirement LSL": ["requirement lsl", "lsl", "lower spec limit", "lower limit", "minimum", "min"],
    "Requirement Nominal": ["requirement nominal", "nominal", "requirement", "dimension", "specified requirement"],
    "Requirement USL": ["requirement usl", "usl", "upper spec limit", "upper limit", "maximum", "max"],
    "Type": ["type", "characteristic type", "dim type"],
    "EZ Fabricating Actual": ["ez fabricating actual", "actual", "actual result", "measured actual", "measurement"],
    "In Spec": ["in spec", "inspec", "pass fail", "pass/fail", "accept", "result"],
    "Tooling Used": ["tooling used", "tooling", "tool used", "gage", "gauge", "inspection tool"],
    "Comments": ["comments", "comment", "notes", "remark", "remarks"],
}

ADMIN_FIELD_PATTERN = re.compile(r"\b(?:PO|PURCHASE\s*ORDER|ORDER\s*(?:NO|NUMBER|#)|INSPECTOR|ITEM\s*(?:NO|NUMBER|#)|REASON)\b", re.I)

# Numbers used as dimensions, optionally prefixed with diameter and/or followed
# by an angular symbol. Requiring a decimal filters many drawing numbers/pages.
DIMENSION_PATTERN = re.compile(r"(?<![A-Za-z0-9])(?P<diam>[Ø⌀])?\s*(?P<num>(?:\d+\.\d+|\.\d+))(?:\s*(?P<angle>°|deg\.?))?", re.I)
TOLERANCE_PATTERN = re.compile(r"\+\s*(?P<plus>\.?\d+(?:\.\d+)?)\s*/?\s*-\s*(?P<minus>\.?\d+(?:\.\d+)?)")
WELD_PATTERN = re.compile(r"(?:\bWELD(?:MENT|ED|ING)?\b|\bFILLET\b|\bSEAM\b|\bBEAD\b|[⌒⏊△])", re.I)
IGNORE_CONTEXT_PATTERN = re.compile(
    r"\b(?:DATE|DWG|DRAWING|REV|REVISION|SHEET|PAGE|PHONE|FAX|ZIP|SCALE|TITLE|CAGE|PART\s*NO|NOTES?)\b",
    re.I,
)
DIMENSIONAL_NOTE_PATTERN = re.compile(r"\b(?:WELD|FILLET|DIAMETER|RADIUS|TOLERANCE|TYP|PLACES?)\b|[Ø⌀°]", re.I)


@dataclass
class Characteristic:
    """A single extracted FAI characteristic and its PDF location."""

    char_number: int
    reference_location: str
    nominal: float
    lsl: float
    usl: float
    type: str
    page_index: int
    rect: tuple[float, float, float, float]
    raw_text: str = ""
    tooling: str = ""
    comments: str = ""
    actual: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)

    def to_row(self) -> dict[str, Any]:
        return {
            "Char Number": self.char_number,
            "Reference Location": self.reference_location,
            "Requirement LSL": self.lsl,
            "Requirement Nominal": self.nominal,
            "Requirement USL": self.usl,
            "Type": self.type,
            "EZ Fabricating Actual": self.actual,
            "Tooling Used": self.tooling,
            "Comments": self.comments,
        }


@dataclass
class SkippedCandidate:
    """A candidate dimension that was rejected during extraction."""

    page_index: int
    raw_text: str
    reason: str
    rect: tuple[float, float, float, float]
    context: str = ""


LAST_EXTRACTION_DEBUG: dict[str, list[dict[str, Any]] | list[SkippedCandidate]] = {"skipped": []}


def get_last_skipped_candidates() -> list[SkippedCandidate]:
    """Return skipped extraction candidates from the last PDF extraction run."""
    return list(LAST_EXTRACTION_DEBUG.get("skipped", []))


def classify_dimension(text: str) -> str:
    """Classify extracted text as LINEAR, diameter, angle, or weld."""
    text = text or ""
    if "°" in text or "º" in text or re.search(r"\bdeg\.?\b", text, re.I):
        return "°"
    if "Ø" in text or "⌀" in text:
        return "Ø"
    if WELD_PATTERN.search(text):
        return "WELD"
    return "LINEAR"


def _decimal_places(value_text: str) -> int:
    if "." not in value_text:
        return 0
    return len(value_text.split(".", 1)[1])


def calculate_tolerance_limits(nominal: float, value_text: str, dim_type: str, context: str = "") -> tuple[float, float]:
    """Calculate inclusive lower/upper tolerance limits for a dimension.

    Explicit bilateral tolerances like +.13 / -.03 win. Otherwise title-block
    defaults are applied: two-place ±0.02, three-place ±0.005, angular ±2.
    """
    tolerance_match = TOLERANCE_PATTERN.search(context or "")
    if tolerance_match:
        plus = float(tolerance_match.group("plus"))
        minus = float(tolerance_match.group("minus"))
        return round(nominal - minus, 6), round(nominal + plus, 6)

    if dim_type == "°":
        tol = TITLE_BLOCK_DEFAULTS["angular"]
    elif _decimal_places(value_text) >= 3:
        tol = TITLE_BLOCK_DEFAULTS["three_place"]
    else:
        tol = TITLE_BLOCK_DEFAULTS["two_place"]
    return round(nominal - tol, 6), round(nominal + tol, 6)


def _reference_location(page: fitz.Page, rect: fitz.Rect, page_index: int) -> str:
    width, height = page.rect.width, page.rect.height
    col = min(4, max(1, int(rect.x0 / max(width / 4, 1)) + 1))
    row = min(4, max(1, int(rect.y0 / max(height / 4, 1)) + 1))
    return f"P{page_index + 1}-R{row}C{col}"


def _nearby_text(page: fitz.Page, rect: fitz.Rect, radius: float = 72) -> str:
    clip = fitz.Rect(rect.x0 - radius, rect.y0 - radius, rect.x1 + radius, rect.y1 + radius) & page.rect
    return page.get_text("text", clip=clip).replace("\n", " ").strip()



def _iter_text_spans(page: fitz.Page) -> Iterable[dict[str, Any]]:
    page_dict = page.get_text("dict")
    for block in page_dict.get("blocks", []):
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                text = span.get("text", "")
                if text.strip():
                    yield span


def _record_skip(skipped: list[SkippedCandidate], page_index: int, raw_text: str, reason: str, rect: fitz.Rect, context: str = "") -> None:
    skipped.append(
        SkippedCandidate(
            page_index=page_index,
            raw_text=raw_text.strip(),
            reason=reason,
            rect=(round(rect.x0, 3), round(rect.y0, 3), round(rect.x1, 3), round(rect.y1, 3)),
            context=" ".join((context or "").split())[:240],
        )
    )


def _noise_reason(text: str, context: str) -> str:
    if not text.strip():
        return "blank candidate"
    combined = f"{text} {context}"
    if IGNORE_CONTEXT_PATTERN.search(combined) and not DIMENSIONAL_NOTE_PATTERN.search(combined):
        return "title block / revision / non-dimensional context"
    compact = re.sub(r"[\s\-().]", "", combined)
    digits = re.sub(r"\D", "", combined)
    if len(digits) >= 7 and not DIMENSIONAL_NOTE_PATTERN.search(combined):
        return "long numeric identifier, phone/date/drawing-number-like text"
    if re.search(r"\b(?:PO|PURCHASE\s*ORDER|ORDER\s*NO|INSPECTOR|ITEM\s*NO|REASON)\b", combined, re.I):
        return "FAI/admin field, not a drawing requirement"
    if re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", compact):
        return "date-like text"
    return ""


def _is_noise(text: str, context: str) -> bool:
    return bool(_noise_reason(text, context))


def _word_rect(word: tuple[Any, ...]) -> fitz.Rect:
    return fitz.Rect(float(word[0]), float(word[1]), float(word[2]), float(word[3]))


def _dimension_candidate_from_words(words: list[tuple[Any, ...]], index: int) -> tuple[str, fitz.Rect]:
    word = words[index]
    text = str(word[4]).strip()
    rect = _word_rect(word)

    if index > 0 and int(words[index - 1][5]) == int(word[5]) and int(words[index - 1][6]) == int(word[6]):
        previous = str(words[index - 1][4]).strip()
        if previous in {"Ø", "⌀"}:
            text = f"{previous}{text}"
            rect |= _word_rect(words[index - 1])

    if index + 1 < len(words) and int(words[index + 1][5]) == int(word[5]) and int(words[index + 1][6]) == int(word[6]):
        next_word = str(words[index + 1][4]).strip()
        if next_word in {"°", "º"} or re.fullmatch(r"deg\.?", next_word, re.I):
            text = f"{text}{next_word}"
            rect |= _word_rect(words[index + 1])

    return text, rect


def _line_text_for_word(words: list[tuple[Any, ...]], index: int) -> str:
    word = words[index]
    block_no = int(word[5])
    line_no = int(word[6])
    return " ".join(str(item[4]).strip() for item in words if int(item[5]) == block_no and int(item[6]) == line_no).strip()


def _looks_like_numeric_candidate(text: str) -> bool:
    return bool(re.search(r"[Ø⌀]?\s*(?:\d+\.\d+|\.\d+)", text))


def extract_pdf_dimensions(pdf_path: str | Path) -> list[Characteristic]:
    """Extract likely dimensional characteristics from a blueprint PDF.

    The function keeps the same simple list return value used by the GUI, and
    stores rejected candidates for the debug report in LAST_EXTRACTION_DEBUG.
    """
    pdf_path = Path(pdf_path)
    characteristics: list[Characteristic] = []
    skipped: list[SkippedCandidate] = []
    seen: set[tuple[int, str, str, int, int]] = set()

    with fitz.open(pdf_path) as doc:
        full_text = "\n".join(page.get_text("text") for page in doc)
        default_comment = "AFTER GALVANIZE" if re.search(r"galvani[sz]e", full_text, re.I) else ""

        for page_index, page in enumerate(doc):
            words = page.get_text("words", sort=True)
            for word_index, word in enumerate(words):
                candidate_text, candidate_rect = _dimension_candidate_from_words(words, word_index)
                if not _looks_like_numeric_candidate(candidate_text):
                    continue

                nearby = _nearby_text(page, candidate_rect)
                line_text = _line_text_for_word(words, word_index)
                if candidate_text.strip().startswith(("+", "-")):
                    _record_skip(skipped, page_index, candidate_text, "explicit tolerance component, not a standalone characteristic", candidate_rect, nearby)
                    continue
                reason = _noise_reason(candidate_text, nearby)
                if reason:
                    _record_skip(skipped, page_index, candidate_text, reason, candidate_rect, nearby)
                    continue

                match = DIMENSION_PATTERN.search(candidate_text)
                if not match:
                    _record_skip(skipped, page_index, candidate_text, "decimal dimension pattern did not match", candidate_rect, nearby)
                    continue

                raw = match.group(0).strip().replace("º", "°")
                number_text = match.group("num")
                try:
                    nominal = float(number_text)
                except ValueError:
                    _record_skip(skipped, page_index, raw, "not numeric after parsing", candidate_rect, nearby)
                    continue
                if nominal <= 0:
                    _record_skip(skipped, page_index, raw, "zero or negative nominal", candidate_rect, nearby)
                    continue

                dim_type = classify_dimension(raw)
                if dim_type == "LINEAR" and WELD_PATTERN.search(nearby):
                    dim_type = "WELD"
                if dim_type == "LINEAR" and nominal > 500:
                    _record_skip(skipped, page_index, raw, "large linear value likely not a part dimension", candidate_rect, nearby)
                    continue

                coarse_key = (page_index, raw, dim_type, round(candidate_rect.x0 / 6), round(candidate_rect.y0 / 6))
                if coarse_key in seen:
                    _record_skip(skipped, page_index, raw, "duplicate candidate at same location", candidate_rect, nearby)
                    continue
                seen.add(coarse_key)

                lsl, usl = calculate_tolerance_limits(nominal, number_text, dim_type, line_text)
                characteristics.append(
                    Characteristic(
                        char_number=len(characteristics) + 1,
                        reference_location=_reference_location(page, candidate_rect, page_index),
                        nominal=nominal,
                        lsl=lsl,
                        usl=usl,
                        type=dim_type,
                        page_index=page_index,
                        rect=(candidate_rect.x0, candidate_rect.y0, candidate_rect.x1, candidate_rect.y1),
                        raw_text=raw,
                        tooling=DEFAULT_TOOLING.get(dim_type, ""),
                        comments=default_comment,
                        metadata={"source": line_text or candidate_text, "nearby": nearby, "drawing_name": pdf_path.stem},
                    )
                )

            # Weld callout MVP: capture one decimal size per weld-text span if it
            # did not already get caught by nearby numeric-word parsing.
            for span in _iter_text_spans(page):
                span_text = span["text"]
                if not WELD_PATTERN.search(span_text):
                    continue
                span_rect = fitz.Rect(span["bbox"])
                nearby = _nearby_text(page, span_rect, radius=96)
                first_num = re.search(r"\d*\.\d+", nearby)
                if not first_num:
                    _record_skip(skipped, page_index, span_text, "weld callout text found but no decimal weld size nearby", span_rect, nearby)
                    continue
                number_text = first_num.group(0)
                nominal = float(number_text)
                key = (page_index, f"WELD-{number_text}", "WELD", round(span_rect.x0 / 6), round(span_rect.y0 / 6))
                if key in seen:
                    _record_skip(skipped, page_index, span_text, "duplicate weld callout candidate", span_rect, nearby)
                    continue
                seen.add(key)
                lsl, usl = calculate_tolerance_limits(nominal, number_text, "WELD", nearby)
                characteristics.append(
                    Characteristic(
                        char_number=len(characteristics) + 1,
                        reference_location=_reference_location(page, span_rect, page_index),
                        nominal=nominal,
                        lsl=lsl,
                        usl=usl,
                        type="WELD",
                        page_index=page_index,
                        rect=(span_rect.x0, span_rect.y0, span_rect.x1, span_rect.y1),
                        raw_text=span_text.strip(),
                        tooling=DEFAULT_TOOLING["WELD"],
                        comments=default_comment,
                        metadata={"source": span_text, "nearby": nearby, "drawing_name": pdf_path.stem},
                    )
                )

    LAST_EXTRACTION_DEBUG["skipped"] = skipped
    return characteristics


def _balloon_position(page: fitz.Page, rect: fitz.Rect, radius: float, occupied: list[fitz.Rect] | None = None) -> fitz.Point:
    occupied = occupied or []
    offsets = [
        (radius * 2.4, -radius * 0.4),
        (-radius * 2.4, -radius * 0.4),
        (radius * 2.2, radius * 1.8),
        (-radius * 2.2, radius * 1.8),
        (radius * 0.4, -radius * 2.4),
        (radius * 0.4, radius * 2.4),
        (radius * 3.4, 0),
        (-radius * 3.4, 0),
    ]
    anchors = [fitz.Point(rect.x1, rect.y0), fitz.Point(rect.x0, rect.y0), fitz.Point(rect.x1, rect.y1), fitz.Point(rect.x0, rect.y1)]
    for anchor in anchors:
        for dx, dy in offsets:
            point = fitz.Point(anchor.x + dx, anchor.y + dy)
            circle = fitz.Rect(point.x - radius, point.y - radius, point.x + radius, point.y + radius)
            if not page.rect.contains(circle):
                continue
            if circle.intersects(rect):
                continue
            if any(circle.intersects(existing) for existing in occupied):
                continue
            return point
    return fitz.Point(
        min(max(rect.x1 + radius, radius), page.rect.width - radius),
        min(max(rect.y0, radius), page.rect.height - radius),
    )


def add_pdf_balloons(pdf_path: str | Path, characteristics: list[Characteristic], output_path: str | Path | None = None) -> Path:
    """Add numbered circular balloons near extracted PDF dimensions."""
    pdf_path = Path(pdf_path)
    output_path = Path(output_path) if output_path else pdf_path.with_name(f"{pdf_path.stem}_BALLOONED.pdf")
    if output_path.resolve() == pdf_path.resolve():
        raise ValueError("Ballooned PDF output cannot overwrite the original PDF.")

    with fitz.open(pdf_path) as doc:
        occupied_by_page: dict[int, list[fitz.Rect]] = {}
        for characteristic in characteristics:
            if characteristic.page_index >= len(doc):
                continue
            page = doc[characteristic.page_index]
            rect = fitz.Rect(characteristic.rect)
            radius = 11
            occupied = occupied_by_page.setdefault(characteristic.page_index, [])
            center = _balloon_position(page, rect, radius, occupied)
            circle = fitz.Rect(center.x - radius, center.y - radius, center.x + radius, center.y + radius)
            page.draw_oval(circle, color=(1, 0, 0), fill=(1, 1, 1), width=1.2)
            page.insert_textbox(
                circle,
                str(characteristic.char_number),
                fontsize=8,
                fontname="helv",
                color=(1, 0, 0),
                align=fitz.TEXT_ALIGN_CENTER,
            )
            page.draw_line(fitz.Point(center.x - radius, center.y), fitz.Point(rect.x0, rect.y0), color=(1, 0, 0), width=0.6)
            occupied.append(circle + (-3, -3, 3, 3))
        doc.save(output_path, garbage=4, deflate=True)
    return output_path


def _normalise_header(header: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(header or "").lower())


def _header_targets() -> dict[str, str]:
    targets: dict[str, str] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        targets[_normalise_header(canonical)] = canonical
        for alias in aliases:
            targets[_normalise_header(alias)] = canonical
    return targets


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    targets = _header_targets()
    best_row = 1
    best_mapping: dict[str, int] = {}
    for row in range(1, min(ws.max_row, 60) + 1):
        mapping: dict[str, int] = {}
        for col in range(1, min(ws.max_column, 80) + 1):
            raw_header = ws.cell(row=row, column=col).value
            value = _normalise_header(raw_header)
            if not value or ADMIN_FIELD_PATTERN.search(str(raw_header or "")):
                continue
            if value in targets:
                mapping.setdefault(targets[value], col)
        if len(mapping) > len(best_mapping):
            best_row, best_mapping = row, mapping

    # A corrected EZ FAB template should already contain most row headers. For
    # very sparse ad-hoc templates, create a simple table instead of writing into
    # random cells. Admin fields like PO/order/inspector/item/reason are never in
    # FAI_HEADERS, so this writer leaves them blank.
    if len(best_mapping) < 4:
        best_row = 1
        best_mapping = {header: index for index, header in enumerate(FAI_HEADERS, start=1)}
        for header, col in best_mapping.items():
            ws.cell(row=best_row, column=col).value = header
    else:
        for header in FAI_HEADERS:
            if header not in best_mapping:
                col = ws.max_column + 1
                ws.cell(row=best_row, column=col).value = header
                best_mapping[header] = col
    return best_row, best_mapping


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.border:
            target.border = copy.copy(source.border)
        if source.fill:
            target.fill = copy.copy(source.fill)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def fill_fai_template(template_path: str | Path, characteristics: list[Characteristic], output_path: str | Path | None = None) -> Path:
    """Fill an Excel FAI template with one row per characteristic."""
    template_path = Path(template_path)
    if output_path is None:
        drawing_name = characteristics[0].metadata.get("drawing_name") if characteristics else template_path.stem
        output_path = template_path.with_name(f"{drawing_name}_FAI.xlsx")
    output_path = Path(output_path)
    if output_path.resolve() == template_path.resolve():
        raise ValueError("FAI Excel output cannot overwrite the original template.")

    wb = load_workbook(template_path)
    ws = wb.active
    header_row, columns = _find_header_row(ws)
    start_row = header_row + 1
    admin_columns = [
        col
        for col in range(1, ws.max_column + 1)
        if ADMIN_FIELD_PATTERN.search(str(ws.cell(row=header_row, column=col).value or ""))
    ]

    for offset, characteristic in enumerate(characteristics):
        row = start_row + offset
        if row > start_row:
            _copy_row_style(ws, start_row, row)
        values = characteristic.to_row()
        for header, value in values.items():
            ws.cell(row=row, column=columns[header]).value = value
        for admin_col in admin_columns:
            ws.cell(row=row, column=admin_col).value = None

        actual_col = columns["EZ Fabricating Actual"]
        lsl_col = columns["Requirement LSL"]
        usl_col = columns["Requirement USL"]
        in_spec_col = columns["In Spec"]
        actual_cell = f"{get_column_letter(actual_col)}{row}"
        lsl_cell = f"{get_column_letter(lsl_col)}{row}"
        usl_cell = f"{get_column_letter(usl_col)}{row}"
        ws.cell(row=row, column=in_spec_col).value = (
            f'=IF({actual_cell}="","",IF(AND({actual_cell}>={lsl_cell},{actual_cell}<={usl_cell}),"X",""))'
        )

    wb.save(output_path)
    return output_path


def write_debug_report(
    pdf_path: str | Path,
    template_path: str | Path,
    characteristics: list[Characteristic],
    output_path: str | Path | None = None,
    skipped_candidates: list[SkippedCandidate] | None = None,
) -> Path:
    """Write a plain-text extraction/debug report beside generated outputs."""
    pdf_path = Path(pdf_path)
    template_path = Path(template_path)
    output_path = Path(output_path) if output_path else pdf_path.with_name("EZ_FAI_DEBUG_REPORT.txt")
    skipped_candidates = skipped_candidates if skipped_candidates is not None else get_last_skipped_candidates()

    lines = [
        "EZ FAI Builder Debug Report",
        "===========================",
        f"PDF file used: {pdf_path}",
        f"Template file used: {template_path}",
        f"Number of extracted characteristics: {len(characteristics)}",
        "",
        "Extracted characteristics:",
    ]
    if not characteristics:
        lines.append("  (none)")
    for characteristic in characteristics:
        lines.extend(
            [
                f"  Char {characteristic.char_number}",
                f"    page: {characteristic.page_index + 1}",
                f"    reference location: {characteristic.reference_location}",
                f"    raw text: {characteristic.raw_text}",
                f"    nominal: {characteristic.nominal}",
                f"    LSL: {characteristic.lsl}",
                f"    USL: {characteristic.usl}",
                f"    type: {characteristic.type}",
                f"    tooling: {characteristic.tooling}",
                f"    comments: {characteristic.comments}",
                f"    PDF coordinate: {tuple(round(v, 3) for v in characteristic.rect)}",
            ]
        )

    lines.extend(["", "Skipped dimension candidates:"])
    if not skipped_candidates:
        lines.append("  (none)")
    for skipped in skipped_candidates:
        lines.extend(
            [
                f"  page {skipped.page_index + 1}: {skipped.raw_text}",
                f"    reason: {skipped.reason}",
                f"    PDF coordinate: {skipped.rect}",
                f"    context: {skipped.context}",
            ]
        )

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output_path


class ReviewTable(ttk.Frame):
    """Editable review table for extracted characteristics."""

    editable_columns = ("nominal", "lsl", "usl", "type", "tooling", "comments")

    def __init__(self, master):
        super().__init__(master)
        self.characteristics: list[Characteristic] = []
        self.columns = ("char", "ref", "nominal", "lsl", "usl", "type", "tooling", "comments", "raw")
        self.tree = ttk.Treeview(self, columns=self.columns, show="headings", height=12)
        headings = {
            "char": "Char #",
            "ref": "Reference Location",
            "nominal": "Nominal",
            "lsl": "LSL",
            "usl": "USL",
            "type": "Type",
            "tooling": "Tooling Used",
            "comments": "Comments",
            "raw": "Raw Text",
        }
        for column, heading in headings.items():
            self.tree.heading(column, text=heading)
            self.tree.column(column, width=110 if column != "comments" else 180)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.bind("<Double-1>", self._begin_edit)

    def load(self, characteristics: list[Characteristic]) -> None:
        self.characteristics = characteristics
        self.refresh()

    def refresh(self) -> None:
        self.tree.delete(*self.tree.get_children())
        for characteristic in self.characteristics:
            self.tree.insert("", tk.END, iid=str(characteristic.char_number), values=(
                characteristic.char_number,
                characteristic.reference_location,
                characteristic.nominal,
                characteristic.lsl,
                characteristic.usl,
                characteristic.type,
                characteristic.tooling,
                characteristic.comments,
                characteristic.raw_text,
            ))

    def delete_selected(self) -> None:
        selected = set(self.tree.selection())
        self.characteristics = [c for c in self.characteristics if str(c.char_number) not in selected]
        for idx, characteristic in enumerate(self.characteristics, start=1):
            characteristic.char_number = idx
        self.refresh()

    def _begin_edit(self, event) -> None:
        item = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        if not item or not column_id:
            return
        column_index = int(column_id.replace("#", "")) - 1
        column_name = self.columns[column_index]
        if column_name not in self.editable_columns:
            return
        bbox = self.tree.bbox(item, column_id)
        if not bbox:
            return
        x, y, width, height = bbox
        value = self.tree.set(item, column_name)
        editor = ttk.Entry(self.tree)
        editor.insert(0, value)
        editor.select_range(0, tk.END)
        editor.focus()
        editor.place(x=x, y=y, width=width, height=height)

        def commit(_event=None):
            new_value = editor.get()
            editor.destroy()
            index = int(item) - 1
            if 0 <= index < len(self.characteristics):
                characteristic = self.characteristics[index]
                if column_name in {"nominal", "lsl", "usl"}:
                    try:
                        new_value = float(new_value)
                    except ValueError:
                        messagebox.showerror("Invalid value", f"{column_name} must be numeric.")
                        return
                setattr(characteristic, column_name, new_value)
                self.refresh()

        editor.bind("<Return>", commit)
        editor.bind("<FocusOut>", commit)
        editor.bind("<Escape>", lambda _event: editor.destroy())


class EZFAIBuilderApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("EZ FAI Builder")
        self.geometry("1100x650")
        self.pdf_path: Path | None = None
        self.template_path: Path | None = None
        self.characteristics: list[Characteristic] = []
        self._build_ui()

    def _build_ui(self) -> None:
        top = ttk.Frame(self, padding=12)
        top.pack(fill=tk.X)

        ttk.Button(top, text="Select PDF", command=self.select_pdf).grid(row=0, column=0, padx=4, pady=4)
        self.pdf_label = ttk.Label(top, text="No PDF selected (drag/drop style: paste or select a file)")
        self.pdf_label.grid(row=0, column=1, sticky="w", padx=4)

        ttk.Button(top, text="Select Excel Template", command=self.select_template).grid(row=1, column=0, padx=4, pady=4)
        self.template_label = ttk.Label(top, text="No Excel template selected")
        self.template_label.grid(row=1, column=1, sticky="w", padx=4)

        ttk.Button(top, text="Extract Dimensions", command=self.extract_dimensions).grid(row=2, column=0, padx=4, pady=8)
        ttk.Button(top, text="Generate Ballooned PDF + FAI Excel", command=self.generate_outputs).grid(row=2, column=1, sticky="w", padx=4, pady=8)
        ttk.Button(top, text="Delete Selected Review Rows", command=self.delete_rows).grid(row=2, column=2, padx=4, pady=8)

        instructions = ttk.Label(
            self,
            text="Review Table: double-click Nominal, LSL, USL, Type, Tooling, or Comments to edit before export.",
            padding=(12, 0),
        )
        instructions.pack(fill=tk.X)

        self.review_table = ReviewTable(self)
        self.review_table.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

        self.status = tk.StringVar(value="Ready")
        ttk.Label(self, textvariable=self.status, relief=tk.SUNKEN, anchor="w").pack(fill=tk.X, side=tk.BOTTOM)

    def select_pdf(self) -> None:
        filename = filedialog.askopenfilename(title="Select blueprint PDF", filetypes=[("PDF files", "*.pdf")])
        if filename:
            self.pdf_path = Path(filename)
            self.pdf_label.config(text=str(self.pdf_path))

    def select_template(self) -> None:
        filename = filedialog.askopenfilename(title="Select FAI Excel template", filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if filename:
            self.template_path = Path(filename)
            self.template_label.config(text=str(self.template_path))

    def extract_dimensions(self) -> None:
        if not self.pdf_path:
            messagebox.showwarning("Missing PDF", "Select a PDF blueprint first.")
            return
        try:
            self.status.set("Extracting dimensions...")
            self.update_idletasks()
            self.characteristics = extract_pdf_dimensions(self.pdf_path)
            for characteristic in self.characteristics:
                characteristic.metadata["drawing_name"] = self.pdf_path.stem
            self.review_table.load(self.characteristics)
            self.status.set(f"Extracted {len(self.characteristics)} candidate characteristics. Review before export.")
        except Exception as exc:  # GUI boundary: show recoverable error to user.
            self.status.set("Extraction failed")
            messagebox.showerror("Extraction failed", str(exc))

    def delete_rows(self) -> None:
        self.review_table.delete_selected()
        self.characteristics = self.review_table.characteristics
        self.status.set(f"{len(self.characteristics)} characteristics remain after delete.")

    def generate_outputs(self) -> None:
        if not self.pdf_path or not self.template_path:
            messagebox.showwarning("Missing files", "Select both a PDF and an Excel template first.")
            return
        if not self.review_table.characteristics:
            messagebox.showwarning("No rows", "Extract and review dimensions before generating outputs.")
            return
        try:
            self.characteristics = self.review_table.characteristics
            for characteristic in self.characteristics:
                characteristic.metadata["drawing_name"] = self.pdf_path.stem
            balloon_path = self.pdf_path.with_name(f"{self.pdf_path.stem}_BALLOONED.pdf")
            fai_path = self.pdf_path.with_name(f"{self.pdf_path.stem}_FAI.xlsx")
            debug_path = self.pdf_path.with_name("EZ_FAI_DEBUG_REPORT.txt")
            add_pdf_balloons(self.pdf_path, self.characteristics, balloon_path)
            fill_fai_template(self.template_path, self.characteristics, fai_path)
            write_debug_report(self.pdf_path, self.template_path, self.characteristics, debug_path)
            self.status.set(f"Created: {balloon_path.name}, {fai_path.name}, and {debug_path.name}")
            messagebox.showinfo("Outputs created", f"Created:\n{balloon_path}\n{fai_path}\n{debug_path}")
        except Exception as exc:  # GUI boundary: show recoverable error to user.
            self.status.set("Output generation failed")
            messagebox.showerror("Output generation failed", str(exc))


def generate_outputs_for_files(pdf_path: str | Path, template_path: str | Path) -> tuple[Path, Path, Path, list[Characteristic]]:
    """Extract, balloon, fill Excel, and write debug report for two local files."""
    pdf_path = Path(pdf_path)
    template_path = Path(template_path)
    characteristics = extract_pdf_dimensions(pdf_path)
    for characteristic in characteristics:
        characteristic.metadata["drawing_name"] = pdf_path.stem
    balloon_path = pdf_path.with_name(f"{pdf_path.stem}_BALLOONED.pdf")
    fai_path = pdf_path.with_name(f"{pdf_path.stem}_FAI.xlsx")
    debug_path = pdf_path.with_name("EZ_FAI_DEBUG_REPORT.txt")
    add_pdf_balloons(pdf_path, characteristics, balloon_path)
    fill_fai_template(template_path, characteristics, fai_path)
    write_debug_report(pdf_path, template_path, characteristics, debug_path)
    return balloon_path, fai_path, debug_path, characteristics


def launch_gui() -> None:
    """Launch the EZ FAI Builder desktop GUI."""
    app = EZFAIBuilderApp()
    app.mainloop()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="EZ FAI Builder local MVP")
    parser.add_argument("pdf", nargs="?", help="Optional PDF blueprint path for batch generation")
    parser.add_argument("template", nargs="?", help="Optional Excel FAI template path for batch generation")
    args = parser.parse_args()
    if args.pdf and args.template:
        ballooned, fai, debug, chars = generate_outputs_for_files(args.pdf, args.template)
        print(f"Extracted {len(chars)} characteristics")
        print(f"Ballooned PDF: {ballooned}")
        print(f"FAI Excel: {fai}")
        print(f"Debug report: {debug}")
    else:
        launch_gui()
